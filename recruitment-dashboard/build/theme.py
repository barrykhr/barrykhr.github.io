# -*- coding: utf-8 -*-
"""Design system for the Zeronorth Recruitment Intelligence workbook."""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"

# ---- palette -------------------------------------------------------------
INK        = "1A1D23"   # primary type
INK_SOFT   = "4B5563"   # secondary type
MUTED      = "8A94A6"   # labels / captions
ACCENT     = "2563EB"   # single primary accent
ACCENT_DK  = "1D4ED8"
ACCENT_TINT= "EFF4FF"
SURFACE    = "FFFFFF"
CANVAS     = "F6F7F9"
LINE       = "E4E7EC"
LINE_SOFT  = "F0F1F4"

GOOD       = "0F9D58"   # green  - strong / positive
GOOD_TINT  = "E7F5EC"
WARN       = "C8860D"   # amber  - attention
WARN_TINT  = "FDF3E2"
RISK       = "D33A3A"   # red    - risk
RISK_TINT  = "FBEAEA"
NEUTRAL    = "6B7280"
NEUTRAL_TINT = "F1F3F6"

# chart series colours (restrained, accent-led)
SERIES = [ACCENT, "7BA4F5", "A9C2F8", "CBD9FB", MUTED, "C3CAD6"]

# ---- helpers -------------------------------------------------------------
def F(size=9, bold=False, color=INK, italic=False):
    return Font(name=FONT, size=size, bold=bold, color=color, italic=italic)

def fill(hexcolor):
    return PatternFill("solid", fgColor=hexcolor)

def side(color=LINE, style="thin"):
    return Side(style=style, color=color)

def box(color=LINE):
    s = side(color)
    return Border(left=s, right=s, top=s, bottom=s)

def bottom_rule(color=LINE, style="thin"):
    return Border(bottom=side(color, style))

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left", vertical="center")
RIGHT  = Alignment(horizontal="right", vertical="center")
LEFTW  = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTERW= Alignment(horizontal="center", vertical="center", wrap_text=True)


def paint(ws, cell_range, fillcolor=None, border=None, font=None, align=None, fmt=None):
    for row in ws[cell_range]:
        for c in row:
            if fillcolor is not None: c.fill = fill(fillcolor)
            if border is not None:    c.border = border
            if font is not None:      c.font = font
            if align is not None:     c.alignment = align
            if fmt is not None:       c.number_format = fmt


def canvas(ws, last_col=40, last_row=200):
    """Give the sheet a flat light ground and hide gridlines."""
    ws.sheet_view.showGridLines = False
    f = fill(CANVAS)
    for r in range(1, last_row + 1):
        for c in range(1, last_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = f
            cell.font = F()


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def heights(ws, spec):
    for r, h in spec.items():
        ws.row_dimensions[r].height = h


def card(ws, r1, c1, r2, c2, fillcolor=SURFACE, edge=LINE):
    """Draw a flat 'card': surface fill + hairline border on the outer ring."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill(fillcolor)
    top, bot = side(edge), side(edge)
    for c in range(c1, c2 + 1):
        ws.cell(row=r1, column=c).border = Border(
            top=top, bottom=ws.cell(row=r1, column=c).border.bottom,
            left=side(edge) if c == c1 else None, right=side(edge) if c == c2 else None)
        ws.cell(row=r2, column=c).border = Border(
            bottom=bot, top=ws.cell(row=r2, column=c).border.top,
            left=side(edge) if c == c1 else None, right=side(edge) if c == c2 else None)
    for r in range(r1, r2 + 1):
        cur = ws.cell(row=r, column=c1)
        cur.border = Border(left=side(edge), top=cur.border.top, bottom=cur.border.bottom,
                            right=cur.border.right)
        cur = ws.cell(row=r, column=c2)
        cur.border = Border(right=side(edge), top=cur.border.top, bottom=cur.border.bottom,
                            left=cur.border.left)


def section_title(ws, row, col, text, sub=None, width=6):
    c = ws.cell(row=row, column=col, value=text)
    c.font = F(11, True, INK)
    c.alignment = LEFT
    if sub:
        s = ws.cell(row=row + 1, column=col, value=sub)
        s.font = F(8, False, MUTED)
        s.alignment = LEFT


def style_chart(ch, title=None, height=7.2, width=12.0, legend=None):
    from openpyxl.chart.text import RichText
    from openpyxl.drawing.text import (RichTextProperties, Paragraph, ParagraphProperties,
                                       CharacterProperties)
    ch.height = height
    ch.width = width
    if title:
        ch.title = title
    ch.style = None
    if legend is None:
        ch.legend = None
    else:
        ch.legend.position = legend
        ch.legend.overlay = False
    # typography
    cp = CharacterProperties(latin=None, sz=800, b=False, solidFill=INK_SOFT)
    try:
        ch.x_axis.txPr = RichText(bodyPr=RichTextProperties(),
                                  p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
        ch.y_axis.txPr = RichText(bodyPr=RichTextProperties(),
                                  p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    except Exception:
        pass
    return ch
