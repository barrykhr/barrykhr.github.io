# -*- coding: utf-8 -*-
"""Layer 3: Recruitment Dashboard presentation sheet."""
from openpyxl.utils import get_column_letter as GL
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from theme import *
from layer1 import (R0, R1, MONEY, PCT, PCT0, INT0, YRS, NUM1, DAYS, B_LEAD, B_KPI, N_LEAD,
                    B_STAGE, B_FUNNEL, B_CITY, B_REQ, B_RECR, B_NOTICE, B_EXP,
                    B_MONTH, B_HIKE)
from layer2 import KPIS

CALC = "'Dashboard Calc'"
KROW = {k[0]: B_KPI + i for i, k in enumerate(KPIS)}


def kpi(name):
    return "={C}!$B${r}".format(C=CALC, r=KROW[name])


def kpi_ref(name):
    return "{C}!$B${r}".format(C=CALC, r=KROW[name])


# card top rows and their four column anchors
CARD_ROWS = [10, 16, 22, 28]
CARD_COLS = [2, 6, 10, 14]      # B, F, J, N

CARDS = [
    # (kpi name, display label, format, value colour, context formula)
    ("Candidates in pipeline", "Candidates in pipeline", INT0, ACCENT,
     '="of "&TEXT({t},"0")&" on file in Master Data-2026"'),
    ("Submitted to client", "Submitted to client", INT0, ACCENT,
     '=TEXT({sr},"0%")&" of the filtered pipeline"'),
    ("In client interview", "In client interview", INT0, ACCENT,
     '=TEXT({ir},"0%")&" of submitted candidates"'),
    ("Offers and hires", "Offers and hires", INT0, GOOD,
     '="Offer stage and beyond"'),

    ("At telephonic screening", "At telephonic screening", INT0, INK,
     '="Earliest live stage"'),
    ("Withdrawn / lost", "Withdrawn / lost", INT0, RISK,
     '="Exited the process"'),
    ("Submission rate", "Submission rate", PCT, GOOD,
     '="Pipeline converted to client submission"'),
    ("Screen-to-interview rate", "Screen to interview", PCT, GOOD,
     '="Submissions converted to interview"'),

    ("Avg current CTC", "Avg current CTC", MONEY, INK,
     '="Across "&TEXT({dc},"0%")&" with CTC recorded"'),
    ("Avg expected CTC", "Avg expected CTC", MONEY, INK,
     '="Mean ask across the filtered pool"'),
    ("Avg expected hike", "Avg expected hike", PCT, WARN,
     '="Uplift sought on current CTC"'),
    ("Highest expected CTC", "Highest expected CTC", MONEY, INK,
     '="Top of the range"'),

    ("Meets experience bar", "Meets experience bar", INT0, GOOD,
     '="At or above "&TEXT(MinExperience,"0")&" years"'),
    ("Joiner-ready", "Joiner-ready", INT0, GOOD,
     '="Notice within "&TEXT(MaxNoticeDays,"0")&" days"'),
    ("Target-location match", "Target-location match", PCT, ACCENT,
     '="Based in or moving to "&TargetLocation'),
    ("Recruiters active", "Recruiters active", INT0, INK,
     '="Working the filtered pipeline"'),
]


def draw_card(ws, r0, c0, label, value_formula, fmt, colour, context_formula):
    c1 = c0 + 3
    card(ws, r0, c0, r0 + 4, c1)
    ws.merge_cells(start_row=r0 + 1, start_column=c0, end_row=r0 + 1, end_column=c1)
    ws.merge_cells(start_row=r0 + 2, start_column=c0, end_row=r0 + 2, end_column=c1)
    ws.merge_cells(start_row=r0 + 3, start_column=c0, end_row=r0 + 3, end_column=c1)
    lab = ws.cell(row=r0 + 1, column=c0, value=label.upper())
    lab.font = F(7.5, True, MUTED)
    lab.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    val = ws.cell(row=r0 + 2, column=c0, value=value_formula)
    val.font = F(20, True, colour)
    val.number_format = fmt
    val.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ctx = ws.cell(row=r0 + 3, column=c0, value=context_formula)
    ctx.font = F(7.5, False, MUTED)
    ctx.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def band_title(ws, row, text, sub):
    t = ws.cell(row=row, column=2, value=text)
    t.font = F(10.5, True, INK)
    t.alignment = Alignment(horizontal="left", vertical="bottom")
    c = ws.cell(row=row + 1, column=2, value=sub)
    c.font = F(8, False, MUTED)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=17)
    ws.merge_cells(start_row=row + 1, start_column=2, end_row=row + 1, end_column=17)


CHART_BANDS = [
    (34, "Pipeline health", "How the filtered pool is distributed across your stages, and where it narrows."),
    (55, "Coverage and supply", "Where candidates are coming from — geography and requisition."),
    (76, "Recruiter and sourcing throughput", "Who is producing pipeline, and how volume has moved over time."),
    (97, "Candidate economics", "What the pool costs, and how ready it is to move."),
    (118, "Experience profile", "Depth of the submitted pool and the uplift it is asking for."),
]
CHART_ROW = {r: r + 3 for r, _t, _s in CHART_BANDS}   # first row of the chart cards

LEAD_TABLE_ROW = 141
MATRIX_ROW = 161


def build_dashboard(wb):
    ws = wb.create_sheet("Recruitment Dashboard", 0)
    canvas(ws, last_col=19, last_row=185)
    widths(ws, {"A": 2.2, "R": 2.2, "S": 2.2})
    for i in range(2, 18):
        ws.column_dimensions[GL(i)].width = 9.6

    hs = {1: 8, 2: 26, 3: 14, 4: 10, 5: 6, 6: 13, 7: 21, 8: 7, 9: 12}
    for r0 in CARD_ROWS:
        hs.update({r0: 6, r0 + 1: 13, r0 + 2: 27, r0 + 3: 13, r0 + 4: 7, r0 + 5: 6})
    for br, _t, _s in CHART_BANDS:
        hs.update({br - 1: 10, br: 16, br + 1: 13, br + 2: 6})
        for k in range(16):
            hs[br + 3 + k] = 14
        hs[br + 19] = 8
    heights(ws, hs)

    # ---- masthead --------------------------------------------------------
    t = ws["B2"]; t.value = "Recruitment Intelligence"
    t.font = F(19, True, INK); t.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("B2:K2")
    s = ws["B3"]
    s.value = ("Zeronorth talent pipeline · every figure below is a formula over "
               "Master Data-2026 and the Submission template — nothing is typed in.")
    s.font = F(8.5, False, MUTED); ws.merge_cells("B3:N3")
    d = ws["O2"]
    d.value = '="Refreshed "&TEXT(TODAY(),"d mmm yyyy")'
    d.font = F(8.5, False, MUTED)
    d.alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells("O2:Q2")
    for c in range(2, 18):
        ws.cell(row=4, column=c).border = Border(bottom=side(LINE))

    # ---- control bar -----------------------------------------------------
    card(ws, 5, 2, 8, 17, SURFACE, LINE)
    controls = [("RECRUITER", 2, 4, "SelRecruiter"), ("LOCATION", 5, 7, "SelCity"),
                ("PIPELINE STAGE", 8, 10, "SelStage"), ("REQUISITION", 11, 13, "SelRequisition")]
    for label, c0, c1, _name in controls:
        ws.merge_cells(start_row=6, start_column=c0, end_row=6, end_column=c1)
        l = ws.cell(row=6, column=c0, value=label)
        l.font = F(7.5, True, MUTED)
        l.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(start_row=7, start_column=c0, end_row=7, end_column=c1)
        v = ws.cell(row=7, column=c0, value="All")
        v.font = F(10, True, ACCENT_DK)
        v.fill = fill(ACCENT_TINT)
        v.border = box("C7D7FB")
        v.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        for c in range(c0, c1 + 1):
            cell = ws.cell(row=7, column=c)
            cell.fill = fill(ACCENT_TINT)
            cell.border = box("C7D7FB")
    ws.merge_cells(start_row=6, start_column=14, end_row=6, end_column=17)
    l = ws.cell(row=6, column=14, value="CURRENT SELECTION")
    l.font = F(7.5, True, MUTED)
    l.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.merge_cells(start_row=7, start_column=14, end_row=7, end_column=17)
    v = ws.cell(row=7, column=14)
    v.value = ('="Showing "&TEXT({a},"0")&" of "&TEXT({b},"0")&" candidates"'
               ).format(a=kpi_ref("Candidates in pipeline"), b=kpi_ref("Total candidates on file"))
    v.font = F(10, True, INK)
    v.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    # ---- KPI cards -------------------------------------------------------
    ctx_vals = dict(t=kpi_ref("Total candidates on file"),
                    sr=kpi_ref("Submission rate"),
                    ir=kpi_ref("Screen-to-interview rate"),
                    dc=kpi_ref("Detail coverage"))
    for i, (name, label, fmt, colour, ctxf) in enumerate(CARDS):
        r0 = CARD_ROWS[i // 4]
        c0 = CARD_COLS[i % 4]
        draw_card(ws, r0, c0, label, kpi(name), fmt, colour, ctxf.format(**ctx_vals))

    # ---- chart band scaffolding -----------------------------------------
    for br, title, sub in CHART_BANDS:
        band_title(ws, br, title, sub)
        card(ws, br + 3, 2, br + 18, 9)
        card(ws, br + 3, 10, br + 18, 17)

    _leaders_table(ws)
    _quality_matrix(ws)
    return ws


LEAD_LAYOUT = [("#", 2, 2, INT0, "c"), ("Candidate", 3, 4, None, "l"),
               ("Company", 5, 6, None, "l"), ("City", 7, 7, None, "l"),
               ("Exp", 8, 8, YRS, "c"), ("Stage", 9, 10, None, "l"),
               ("Sub-status", 11, 12, None, "l"), ("CCTC", 13, 13, MONEY, "c"),
               ("ECTC", 14, 14, MONEY, "c"), ("Hike", 15, 15, PCT, "c"),
               ("Availability", 16, 16, None, "c"), ("Recruiter", 17, 17, None, "l")]


def _table_header(ws, row, layout):
    for label, c0, c1, _fmt, al in layout:
        if c1 > c0:
            ws.merge_cells(start_row=row, start_column=c0, end_row=row, end_column=c1)
        for c in range(c0, c1 + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = fill(NEUTRAL_TINT)
            cell.border = Border(bottom=side(LINE), top=side(LINE))
        h = ws.cell(row=row, column=c0, value=label.upper())
        h.font = F(7.5, True, INK_SOFT)
        h.alignment = Alignment(horizontal={"l": "left", "c": "center"}[al],
                                vertical="center", indent=1 if al == "l" else 0)


def _leaders_table(ws):
    r = LEAD_TABLE_ROW
    heights(ws, {r - 3: 12, r - 2: 16, r - 1: 13, r: 20})
    band_title(ws, r - 2, "Pipeline leaders",
               "Top 12 by furthest stage reached, then most recent activity. "
               "Recalculates the moment a stage changes — no manual sorting.")
    card(ws, r, 2, r + 13, 17)
    _table_header(ws, r, LEAD_LAYOUT)
    src_cols = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]
    for i in range(N_LEAD):
        rr = r + 1 + i
        ws.row_dimensions[rr].height = 17
        for (label, c0, c1, fmt, al), sc in zip(LEAD_LAYOUT, src_cols):
            if c1 > c0:
                ws.merge_cells(start_row=rr, start_column=c0, end_row=rr, end_column=c1)
            for c in range(c0, c1 + 1):
                cell = ws.cell(row=rr, column=c)
                cell.fill = fill(SURFACE if i % 2 == 0 else "FAFBFC")
                cell.border = bottom_rule(LINE_SOFT)
            cell = ws.cell(row=rr, column=c0)
            cell.value = "={C}!${sc}${sr}".format(C=CALC, sc=sc, sr=B_LEAD + i)
            cell.font = F(8.5, i == 0, INK if label in ("Candidate", "#") else INK_SOFT)
            if fmt:
                cell.number_format = fmt
            cell.alignment = Alignment(horizontal={"l": "left", "c": "center"}[al],
                                       vertical="center", indent=1 if al == "l" else 0)
    rng = "I{a}:J{b}".format(a=r + 1, b=r + N_LEAD)
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=['ISNUMBER(SEARCH("Interview",$I{a}))'.format(a=r + 1)],
        fill=fill(GOOD_TINT), font=F(8.5, True, GOOD), stopIfTrue=False))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=['ISNUMBER(SEARCH("Submission",$I{a}))'.format(a=r + 1)],
        fill=fill(ACCENT_TINT), font=F(8.5, False, ACCENT_DK), stopIfTrue=False))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=['ISNUMBER(SEARCH("Withdrawn",$I{a}))'.format(a=r + 1)],
        fill=fill(RISK_TINT), font=F(8.5, False, RISK), stopIfTrue=False))


MATRIX_LAYOUT = [("#", 2, 2, INT0, "c"), ("Candidate", 3, 5, None, "l"),
                 ("City", 6, 7, None, "l"), ("Experience", 8, 8, NUM1, "c"),
                 ("Current CTC", 9, 9, NUM1, "c"), ("Expected CTC", 10, 10, NUM1, "c"),
                 ("Hike %", 11, 11, PCT0, "c"), ("Notice days", 12, 12, INT0, "c"),
                 ("Stage reached", 13, 13, INT0, "c"), ("Loc match", 14, 14, INT0, "c"),
                 ("Current company", 15, 17, None, "l")]


def _quality_matrix(ws):
    r = MATRIX_ROW
    heights(ws, {r - 3: 12, r - 2: 16, r - 1: 13, r: 26})
    band_title(ws, r - 2, "Candidate comparison matrix",
               "The same 12 candidates across every recorded dimension. Green is strong, "
               "amber is a caveat, red is a risk — scaled live against the rest of the pool.")
    card(ws, r, 2, r + 13, 17)
    _table_header(ws, r, MATRIX_LAYOUT)
    src = ["A", "B", "D", "E", "H", "I", "J", "K", "F", "M", "C"]
    for i in range(N_LEAD):
        rr = r + 1 + i
        ws.row_dimensions[rr].height = 17
        for (label, c0, c1, fmt, al), sc in zip(MATRIX_LAYOUT, src):
            if c1 > c0:
                ws.merge_cells(start_row=rr, start_column=c0, end_row=rr, end_column=c1)
            for c in range(c0, c1 + 1):
                cell = ws.cell(row=rr, column=c)
                cell.fill = fill(SURFACE if i % 2 == 0 else "FAFBFC")
                cell.border = bottom_rule(LINE_SOFT)
            cell = ws.cell(row=rr, column=c0)
            if label == "Notice days":
                cell.value = ('=IF({C}!$O${sr}="","",INDEX(\'Dashboard Calc\'!$S${a}:$S${b},'
                              '{C}!$O${sr}))').format(C=CALC, sr=B_LEAD + i, a=R0, b=R1)
            elif label == "Stage reached":
                cell.value = ('=IF({C}!$O${sr}="","",INDEX(\'Dashboard Calc\'!$J${a}:$J${b},'
                              '{C}!$O${sr}))').format(C=CALC, sr=B_LEAD + i, a=R0, b=R1)
            else:
                cell.value = "={C}!${sc}${sr}".format(C=CALC, sc=sc, sr=B_LEAD + i)
            cell.font = F(8.5, False, INK if label == "Candidate" else INK_SOFT)
            if fmt:
                cell.number_format = fmt
            cell.alignment = Alignment(horizontal={"l": "left", "c": "center"}[al],
                                       vertical="center", indent=1 if al == "l" else 0)
    a, b = r + 1, r + N_LEAD
    green_hi = ColorScaleRule(start_type="min", start_color="FFFFFF",
                              end_type="max", end_color="BFE3CC")
    amber_hi = ColorScaleRule(start_type="min", start_color="FFFFFF",
                              end_type="max", end_color="F6DFB0")
    red_hi   = ColorScaleRule(start_type="min", start_color="FFFFFF",
                              end_type="max", end_color="F3C4C4")
    green_lo = ColorScaleRule(start_type="min", start_color="BFE3CC",
                              end_type="max", end_color="FFFFFF")
    ws.conditional_formatting.add("H%d:H%d" % (a, b), green_hi)   # experience
    ws.conditional_formatting.add("I%d:I%d" % (a, b), amber_hi)   # current CTC
    ws.conditional_formatting.add("J%d:J%d" % (a, b), amber_hi)   # expected CTC
    ws.conditional_formatting.add("K%d:K%d" % (a, b), red_hi)     # hike ask
    ws.conditional_formatting.add("L%d:L%d" % (a, b), green_lo)   # notice days
    ws.conditional_formatting.add("M%d:M%d" % (a, b), green_hi)   # stage reached
    ws.conditional_formatting.add("N%d:N%d" % (a, b), CellIsRule(
        operator="equal", formula=["1"], fill=fill(GOOD_TINT), font=F(8.5, True, GOOD)))
    ws.conditional_formatting.add("N%d:N%d" % (a, b), CellIsRule(
        operator="equal", formula=["0"], fill=fill(NEUTRAL_TINT), font=F(8.5, False, MUTED)))
    note = ws.cell(row=r + 15, column=2)
    note.value = ("Notice days: 0 = immediate, –1 = not recorded.  Loc match: 1 = based in or "
                  "relocating to the target location.  Stage reached: 1 screening → 5 hired.")
    note.font = F(7.5, False, MUTED)
    ws.merge_cells(start_row=r + 15, start_column=2, end_row=r + 15, end_column=17)
