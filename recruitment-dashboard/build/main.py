# -*- coding: utf-8 -*-
import sys, warnings
warnings.filterwarnings("ignore")
import openpyxl
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter as GL

from theme import *
from layer1 import build_reference, build_calc, R0, R1, SRC, LOC_ROW0, LOC_ROW1, PARAM_ROW
import layer2, layer3, layer4, layer5

SRCFILE, OUTFILE = sys.argv[1], sys.argv[2]

wb = openpyxl.load_workbook(SRCFILE)
existing = wb.sheetnames[:]
print("existing sheets:", existing)

# --------------------------------------------------------------------------
# 0 · light, non-destructive preparation of the source sheets
# --------------------------------------------------------------------------
md = wb["Master Data-2026"]
if not md["M1"].value:
    md["M1"] = "Additional Notes"          # header was blank; needed for the table + lookups
for c in range(1, 14):
    cell = md.cell(row=1, column=c)
    cell.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
    cell.fill = fill(INK)
    cell.alignment = CENTERW
md.row_dimensions[1].height = 26
md.freeze_panes = "B2"
last_md = 2
for r in range(2, md.max_row + 1):
    if md.cell(row=r, column=4).value not in (None, ""):
        last_md = r
try:
    md.add_table(Table(displayName="tblMaster2026",
                       ref="A1:M{0}".format(max(last_md, 2)),
                       tableStyleInfo=TableStyleInfo(name="TableStyleLight9",
                                                     showRowStripes=True)))
except Exception as e:
    print("table skipped:", e)

# --------------------------------------------------------------------------
# 1 · build the new layers
# --------------------------------------------------------------------------
ws_ref  = build_reference(wb)
ws_calc = build_calc(wb)
layer2.build_aggregates(ws_calc)
layer2.build_ranked_and_kpis(ws_calc)
ws_dash = layer3.build_dashboard(wb)
ws_cli  = layer5.build_client_view(wb)
ws_rec  = layer5.build_recruiter_view(wb)
ws_det  = layer5.build_candidate_detail(wb)
layer4.build_charts(wb, ws_dash, ws_calc)
layer4.build_client_charts(wb, ws_cli, ws_calc)

# --------------------------------------------------------------------------
# 2 · named ranges — the contract between the layers
# --------------------------------------------------------------------------
NAMES = {
    "StageNameRef":     "Reference!$B$7:$B$13",
    "StageOrderRef":    "Reference!$C$7:$C$13",
    "LocPattern":       "Reference!$B${L0}:$B${L1}".format(L0=LOC_ROW0, L1=LOC_ROW1),
    "LocCity":          "Reference!$C${L0}:$C${L1}".format(L0=LOC_ROW0, L1=LOC_ROW1),
    "LocIdx":           "Reference!$D${L0}:$D${L1}".format(L0=LOC_ROW0, L1=LOC_ROW1),
    "TargetLocation":   "Reference!$C$" + str(PARAM_ROW + 1),
    "MinExperience":    "Reference!$C$" + str(PARAM_ROW + 2),
    "SeniorExperience": "Reference!$C$" + str(PARAM_ROW + 3),
    "MaxNoticeDays":    "Reference!$C$" + str(PARAM_ROW + 4),
    "SelRecruiter":     "'Recruitment Dashboard'!$B$7",
    "SelCity":          "'Recruitment Dashboard'!$E$7",
    "SelStage":         "'Recruitment Dashboard'!$H$7",
    "SelRequisition":   "'Recruitment Dashboard'!$K$7",
    "SelCandidate":     "'Candidate Detail'!$B$7",
    "ListRecruiter":    "OFFSET('Dashboard Calc'!$BH$6,0,0,MAX(1,SUMPRODUCT(('Dashboard Calc'!$BH$6:$BH$21<>\"\")*1)),1)",
    "ListCity":         "OFFSET('Dashboard Calc'!$BI$6,0,0,MAX(1,SUMPRODUCT(('Dashboard Calc'!$BI$6:$BI$21<>\"\")*1)),1)",
    "ListStage":        "OFFSET('Dashboard Calc'!$BJ$6,0,0,MAX(1,SUMPRODUCT(('Dashboard Calc'!$BJ$6:$BJ$21<>\"\")*1)),1)",
    "ListRequisition":  "OFFSET('Dashboard Calc'!$BK$6,0,0,MAX(1,SUMPRODUCT(('Dashboard Calc'!$BK$6:$BK$21<>\"\")*1)),1)",
    "ListCandidate":    "OFFSET('Dashboard Calc'!$BL$6,0,0,MAX(1,SUMPRODUCT(('Dashboard Calc'!$BL$6:$BL$" + str(R1) + "<>\"\")*1)),1)",
}
for nm, ref in NAMES.items():
    wb.defined_names.add(DefinedName(nm, attr_text=ref))

# --------------------------------------------------------------------------
# 3 · data validation
# --------------------------------------------------------------------------
def dv(ws, ref, formula, title, msg):
    d = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    d.promptTitle, d.prompt = title, msg
    d.showInputMessage = True
    d.errorTitle = "Value not in the list"
    d.error = msg
    d.showErrorMessage = True
    ws.add_data_validation(d)
    d.add(ref)

dv(ws_dash, "B7", "=ListRecruiter", "Recruiter", "Pick a recruiter, or All.")
dv(ws_dash, "E7", "=ListCity",      "Location",  "Pick a normalised city, or All.")
dv(ws_dash, "H7", "=ListStage",     "Stage",     "Pick a pipeline stage, or All.")
dv(ws_dash, "K7", "=ListRequisition","Requisition","Pick a requisition, or All.")
dv(ws_det,  "B7", "=ListCandidate", "Candidate", "Pick any candidate in Master Data-2026.")
md.data_validations.dataValidation = [
    d for d in md.data_validations.dataValidation if str(d.sqref) != "H2:H54"]
dv(md, "H2:H400", "=StageNameRef", "Screening status",
   "Use one of the controlled stage values from the Reference sheet.")

num = DataValidation(type="decimal", operator="between", formula1="0", formula2="60",
                     allow_blank=False)
num.errorTitle = "Out of range"
num.error = "Experience bars must be between 0 and 60 years."
num.showErrorMessage = True
ws_ref.add_data_validation(num); num.add("C{0}:C{1}".format(PARAM_ROW + 2, PARAM_ROW + 3))
nd = DataValidation(type="whole", operator="between", formula1="0", formula2="365",
                    allow_blank=False)
nd.errorTitle = "Out of range"
nd.error = "Notice window must be between 0 and 365 days."
nd.showErrorMessage = True
ws_ref.add_data_validation(nd); nd.add("C{0}".format(PARAM_ROW + 4))

# --------------------------------------------------------------------------
# 4 · sheet order, tab colours, print setup
# --------------------------------------------------------------------------
order = ["Recruitment Dashboard", "Client View", "Recruiter View", "Candidate Detail"] \
        + existing + ["Dashboard Calc", "Reference"]
wb._sheets = [wb[n] for n in order if n in wb.sheetnames]
for n, col in (("Recruitment Dashboard", "2563EB"), ("Client View", "0F9D58"),
               ("Recruiter View", "1A1D23"), ("Candidate Detail", "6B7280"),
               ("Dashboard Calc", "C3CAD6"), ("Reference", "C3CAD6")):
    wb[n].sheet_properties.tabColor = col
for n in existing:
    wb[n].sheet_properties.tabColor = "E4E7EC"

for n in ("Recruitment Dashboard", "Client View", "Candidate Detail"):
    s = wb[n]
    s.page_setup.orientation = "portrait"
    s.page_setup.fitToWidth = 1
    s.page_setup.fitToHeight = 0
    s.sheet_properties.pageSetUpPr.fitToPage = True
wb["Recruiter View"].page_setup.orientation = "landscape"

wb.calculation.fullCalcOnLoad = True
wb.active = 0
wb.save(OUTFILE)
print("saved", OUTFILE)
