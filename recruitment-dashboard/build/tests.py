# -*- coding: utf-8 -*-
"""Mutate only SOURCE data (or a dashboard filter cell), recalc, and check the
dashboard followed. Nothing in the calc/presentation layers is ever touched."""
import os, sys, json, shutil, subprocess, warnings
warnings.filterwarnings("ignore")
import openpyxl
from probe import probe

EVAL = os.path.join(os.path.dirname(os.path.abspath(__file__)), "evalwb.py")
BUILT = sys.argv[1]
WORK = os.path.dirname(os.path.abspath(BUILT))


def find_row(ws, name):
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=4).value
        if v and str(v).strip() == name:
            return r
    raise SystemExit("candidate not found: " + name)


def m_stage(wb):
    ws = wb["Master Data-2026"]
    r = find_row(ws, "Porselvan subramanian")
    ws.cell(row=r, column=8).value = "ZN Interview"
    return "Porselvan subramanian: Telephonic Screening -> ZN Interview"


def m_ctc(wb):
    ws = wb["Master Data-2026"]
    r = find_row(ws, "Sibiraj Munirathinam")
    ws.cell(row=r, column=11).value = "30 LPA"
    return "Sibiraj Munirathinam expected CTC: 12 LPA -> 30 LPA"


def m_add(wb):
    ws = wb["Master Data-2026"]
    r = 55
    vals = ["Sarath", ws.cell(row=2, column=2).value, "August", "Meera Krishnan",
            "9812345678", "Pune", "Full Stack Lead ", "ZN Offer", "Offer released",
            "22 LPA", "34 LPA", "Immediate", "Added by dynamic-behaviour test"]
    for i, v in enumerate(vals):
        ws.cell(row=r, column=1 + i).value = v
    sub = wb["Submission template"]
    srow = 43
    s = [26.0, "Kivitronics", ws.cell(row=2, column=2).value, "Meera Krishnan",
         "9812345678", "meera.k@example.com", "meera.k@example.com", "B.Tech",
         "11 Years", "React 9Y + Node 8Y", "Thoughtworks", "FullStack Lead (React,Node)",
         "Principal Engineer", "Pune", "Chennai", "22 LPA", "34 LPA", "Immediate",
         "Career Growth"]
    for i, v in enumerate(s):
        sub.cell(row=srow, column=1 + i).value = v
    return "New candidate Meera Krishnan (Pune, ZN Offer, 11 yrs) appended to both sources"


def m_location(wb):
    ws = wb["Master Data-2026"]
    for name in ("Kishore M", "Suriya", "Rishi Kumar"):
        ws.cell(row=find_row(ws, name), column=6).value = "Hyderabad"
    return "Three Chennai candidates relocated to Hyderabad"


def m_notice(wb):
    ws = wb["Master Data-2026"]
    r = find_row(ws, "Yuva Charan Reddy Durgam")
    ws.cell(row=r, column=12).value = "Immediate"
    r2 = find_row(ws, "Saravana Ganesh S")
    ws.cell(row=r2, column=12).value = "90 days"
    return "Notice periods changed: Yuva Charan -> Immediate, Saravana Ganesh -> 90 days"


def m_filter(wb):
    wb["Recruitment Dashboard"]["E7"] = "Chennai"
    wb["Recruitment Dashboard"]["H7"] = "ZN Submission"
    return "Dashboard filters set to City=Chennai, Stage=ZN Submission"


def m_delete(wb):
    ws = wb["Master Data-2026"]
    r = find_row(ws, "Govardhan")
    for c in range(1, 14):
        ws.cell(row=r, column=c).value = None
    return "Candidate Govardhan removed from the source"


TESTS = [("T1_stage_change", m_stage), ("T2_ctc_change", m_ctc),
         ("T3_add_candidate", m_add), ("T4_location_change", m_location),
         ("T5_notice_change", m_notice), ("T6_filter", m_filter),
         ("T7_delete_candidate", m_delete)]


def run(name, mutate):
    path = os.path.join(WORK, "test_%s.xlsx" % name)
    shutil.copy(BUILT, path)
    wb = openpyxl.load_workbook(path)
    desc = mutate(wb)
    wb.save(path)
    out = os.path.join(WORK, "calc_%s.xlsx" % name)
    res = subprocess.run([sys.executable, EVAL, path, out],
                         capture_output=True, text=True)
    line = [l for l in res.stdout.splitlines() if l.startswith("cells written")]
    return {"change": desc, "engine": line[0] if line else res.stdout[-300:],
            "state": probe(out)}


if __name__ == "__main__":
    only = sys.argv[2:] if len(sys.argv) > 2 else None
    out = {}
    for name, fn in TESTS:
        if only and name not in only:
            continue
        sys.stderr.write("running %s\n" % name); sys.stderr.flush()
        out[name] = run(name, fn)
        tag = "_".join(only) if only else "all"
        with open(os.path.join(WORK, "results_%s.json" % tag), "w") as f:
            json.dump(out, f, indent=1, default=str)
    print("done")
