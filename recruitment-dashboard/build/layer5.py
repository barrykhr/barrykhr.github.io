# -*- coding: utf-8 -*-
"""Layer 5: Client View, Recruiter View, Candidate Detail."""
from openpyxl.utils import get_column_letter as GL
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from theme import *
from layer1 import (R0, R1, MONEY, PCT, PCT0, INT0, YRS, NUM1, SRC,
                    B_LEAD, B_KPI, B_CITY, B_RECR, B_REQ, B_EXP, B_NOTICE,
                    B_FUNNEL, B_STAGE, N_LEAD, FUNNEL, STAGES)
from layer2 import KPIS
from layer3 import CALC, KROW, kpi, kpi_ref, draw_card, band_title, _table_header

CD = "'Candidate Detail'"


# ==========================================================================
# CLIENT VIEW
# ==========================================================================
CV_CARDS = [
    ("Total candidates on file", "Candidates reviewed", INT0, ACCENT),
    ("Submitted to client", "Submitted for review", INT0, ACCENT),
    ("In client interview", "Progressed to interview", INT0, GOOD),
    ("Avg experience", "Average experience", YRS, INK),
    ("Avg expected CTC", "Average expected CTC", MONEY, INK),
    ("Target-location match", "In target location", PCT, ACCENT),
]

CV_TABLE = [("#", 2, 2, INT0, "c"), ("Candidate", 3, 5, None, "l"),
            ("Current company", 6, 8, None, "l"), ("Experience", 9, 9, YRS, "c"),
            ("Location", 10, 11, None, "l"), ("Expected CTC", 12, 12, MONEY, "c"),
            ("Availability", 13, 14, None, "c"), ("Stage reached", 15, 17, None, "l")]


def build_client_view(wb):
    ws = wb.create_sheet("Client View")
    canvas(ws, last_col=19, last_row=110)
    widths(ws, {"A": 2.2, "R": 2.2})
    for i in range(2, 18):
        ws.column_dimensions[GL(i)].width = 9.6
    heights(ws, {1: 10, 2: 28, 3: 15, 4: 12, 5: 8})

    ws["B2"] = "Candidate Shortlist & Market Read"
    ws["B2"].font = F(19, True, INK); ws.merge_cells("B2:L2")
    ws["B3"] = ('="Prepared for the hiring manager · "&TEXT(TODAY(),"d mmmm yyyy")'
                '&" · "&TEXT({t},"0")&" candidates reviewed"'
                ).format(t=kpi_ref("Total candidates on file"))
    ws["B3"].font = F(9, False, MUTED); ws.merge_cells("B3:L3")
    for c in range(2, 18):
        ws.cell(row=4, column=c).border = Border(bottom=side(LINE))

    # ---- executive summary ----------------------------------------------
    band_title(ws, 6, "Executive summary",
               "Written from the live pipeline — it rewrites itself as the data changes.")
    heights(ws, {6: 16, 7: 13, 8: 6})
    card(ws, 9, 2, 16, 17)
    paras = [
        ('="Zeronorth has screened "&TEXT({tot},"0")&" candidates across "&TEXT({req},"0")'
         '&" live requisitions. "&TEXT({sub},"0")&" have been submitted for client review — '
         '"&TEXT({sr},"0%")&" of the pipeline — and "&TEXT({intv},"0")&" have reached '
         'interview stage."'),
        ('="The submitted pool averages "&TEXT({exp},"0.0")&" years of experience against an '
         'average expected CTC of "&TEXT({ectc},"0.0")&" LPA, an uplift of "&TEXT({hike},"0%")'
         '&" on current package. "&TEXT({bar},"0")&" candidates meet or exceed the '
         '"&TEXT(MinExperience,"0")&"-year experience bar."'),
        ('="Supply is concentrated in "&{city}&". "&TEXT({loc},"0%")&" of candidates are based '
         'in or willing to relocate to "&TargetLocation&", and "&TEXT({ready},"0")&" can join '
         'within "&TEXT(MaxNoticeDays,"0")&" days."'),
    ]
    vals = dict(tot=kpi_ref("Total candidates on file"), req=kpi_ref("Open requisitions"),
                sub=kpi_ref("Submitted to client"), sr=kpi_ref("Submission rate"),
                intv=kpi_ref("In client interview"), exp=kpi_ref("Avg experience"),
                ectc=kpi_ref("Avg expected CTC"), hike=kpi_ref("Avg expected hike"),
                bar=kpi_ref("Meets experience bar"), city=kpi_ref("Strongest city"),
                loc=kpi_ref("Target-location match"), ready=kpi_ref("Joiner-ready"))
    for i, p in enumerate(paras):
        r = 10 + i * 2
        ws.merge_cells(start_row=r, start_column=3, end_row=r + 1, end_column=16)
        c = ws.cell(row=r, column=3, value=p.format(**vals))
        c.font = F(9.5, False, INK_SOFT)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 15
        ws.row_dimensions[r + 1].height = 15

    # ---- KPI strip -------------------------------------------------------
    heights(ws, {17: 12, 18: 6, 19: 13, 20: 26, 21: 12, 22: 7, 23: 6,
                 24: 6, 25: 13, 26: 26, 27: 12, 28: 7, 29: 12})
    for i, (name, label, fmt, colour) in enumerate(CV_CARDS):
        r0 = 18 if i < 3 else 24
        c0 = [2, 8, 13][i % 3] if i < 3 else [2, 8, 13][i % 3]
        c1 = [7, 12, 17][i % 3]
        card(ws, r0, c0, r0 + 4, c1)
        ws.merge_cells(start_row=r0 + 1, start_column=c0, end_row=r0 + 1, end_column=c1)
        ws.merge_cells(start_row=r0 + 2, start_column=c0, end_row=r0 + 2, end_column=c1)
        l = ws.cell(row=r0 + 1, column=c0, value=label.upper())
        l.font = F(7.5, True, MUTED)
        l.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        v = ws.cell(row=r0 + 2, column=c0, value=kpi(name))
        v.font = F(20, True, colour); v.number_format = fmt
        v.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ---- recommended candidates -----------------------------------------
    band_title(ws, 31, "Recommended candidates",
               "The five furthest through the process, ranked automatically.")
    heights(ws, {31: 16, 32: 13, 33: 6, 34: 22})
    card(ws, 34, 2, 40, 17)
    _table_header(ws, 34, CV_TABLE)
    src = ["A", "B", "C", "E", "D", "I", "K", "F"]
    for i in range(5):
        rr = 35 + i
        ws.row_dimensions[rr].height = 19
        for (label, c0, c1, fmt, al), sc in zip(CV_TABLE, src):
            if c1 > c0:
                ws.merge_cells(start_row=rr, start_column=c0, end_row=rr, end_column=c1)
            for c in range(c0, c1 + 1):
                cell = ws.cell(row=rr, column=c)
                cell.fill = fill(SURFACE if i % 2 == 0 else "FAFBFC")
                cell.border = bottom_rule(LINE_SOFT)
            cell = ws.cell(row=rr, column=c0)
            cell.value = "={C}!${sc}${sr}".format(C=CALC, sc=sc, sr=B_LEAD + i)
            cell.font = F(9.5, label == "Candidate", INK if label in ("Candidate", "#")
                          else INK_SOFT)
            if fmt:
                cell.number_format = fmt
            cell.alignment = Alignment(horizontal={"l": "left", "c": "center"}[al],
                                       vertical="center", indent=1 if al == "l" else 0)

    # ---- market insights -------------------------------------------------
    band_title(ws, 43, "Hiring market read",
               "Where this pool is strong, and where it will cost you.")
    heights(ws, {43: 16, 44: 13, 45: 6})
    card(ws, 46, 2, 60, 9)
    card(ws, 46, 10, 60, 17)
    left = [
        ("Deepest supply", '={c}', None, "Most candidates available"),
        ("Requisitions in play", kpi("Open requisitions"), INT0, "Distinct roles being worked"),
        ("Recruiters engaged", kpi("Recruiters active"), INT0, "Working this pipeline"),
        ("Target-location match", kpi("Target-location match"), PCT, "Based in or moving to target"),
        ("Immediately available", kpi("Immediate joiners"), INT0, "Can start now"),
        ("Joiner-ready", kpi("Joiner-ready"), INT0, "Within the acceptable notice window"),
    ]
    right = [
        ("Average experience", kpi("Avg experience"), YRS, "Across profiles with detail on file"),
        ("Most experienced", kpi("Most experienced"), YRS, "Deepest profile reviewed"),
        ("Senior / lead calibre", kpi("Senior / lead calibre"), INT0, "At or above the senior bar"),
        ("Average current CTC", kpi("Avg current CTC"), MONEY, "What the pool earns today"),
        ("Average expected CTC", kpi("Avg expected CTC"), MONEY, "What the pool is asking"),
        ("Average uplift sought", kpi("Avg expected hike"), PCT, "Premium over current package"),
    ]
    for col_items, c0, c1, heading in ((left, 2, 9, "Supply and coverage"),
                                       (right, 10, 17, "Depth and cost")):
        h = ws.cell(row=47, column=c0 + 1, value=heading.upper())
        h.font = F(7.5, True, MUTED)
        for i, (label, formula, fmt, note) in enumerate(col_items):
            rr = 49 + i * 2
            ws.row_dimensions[rr].height = 15
            ws.row_dimensions[rr + 1].height = 12
            ws.merge_cells(start_row=rr, start_column=c0 + 1, end_row=rr, end_column=c1 - 3)
            a = ws.cell(row=rr, column=c0 + 1, value=label)
            a.font = F(9.5, False, INK)
            a.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(start_row=rr, start_column=c1 - 2, end_row=rr, end_column=c1)
            b = ws.cell(row=rr, column=c1 - 2,
                        value=formula.format(c=kpi_ref("Strongest city")))
            b.font = F(11, True, ACCENT)
            if fmt:
                b.number_format = fmt
            b.alignment = Alignment(horizontal="right", vertical="center", indent=1)
            ws.merge_cells(start_row=rr + 1, start_column=c0 + 1, end_row=rr + 1, end_column=c1)
            n = ws.cell(row=rr + 1, column=c0 + 1, value=note)
            n.font = F(7.5, False, MUTED)
            for c in range(c0 + 1, c1 + 1):
                ws.cell(row=rr + 1, column=c).border = Border(bottom=side(LINE_SOFT))
    band_title(ws, 63, "Pipeline and geography at a glance",
               "The same two views the recruiting team works from.")
    heights(ws, {62: 12, 63: 16, 64: 13, 65: 6})
    card(ws, 66, 2, 82, 9)
    card(ws, 66, 10, 82, 17)
    for k in range(66, 83):
        ws.row_dimensions[k].height = 14

    ws["B85"] = ("Figures cover candidates recorded in Master Data-2026. Experience, company "
                 "and designation are drawn from the Submission template detail block, so they "
                 "are available for submitted candidates.")
    ws["B85"].font = F(7.5, False, MUTED)
    ws.merge_cells("B85:Q85")
    return ws


# ==========================================================================
# RECRUITER VIEW
# ==========================================================================
RV_COLS = [
    ("Candidate",      "B",  22, None,  "l"),
    ("Recruiter",      "C",  12, None,  "l"),
    ("Sourced",        "D",  11, "dd-mmm-yy", "c"),
    ("Requisition",    "F",  22, None,  "l"),
    ("City",           "H",  13, None,  "l"),
    ("Stage",          "I",  19, None,  "l"),
    ("Stage #",        "J",   8, INT0,  "c"),
    ("Sub-status",     "K",  30, None,  "l"),
    ("Exp",            "W",   9, NUM1,  "c"),
    ("Current company","Y",  22, None,  "l"),
    ("Designation",    "Z",  22, None,  "l"),
    ("CCTC",           "M",  10, NUM1,  "c"),
    ("ECTC",           "O",  10, NUM1,  "c"),
    ("Hike",           "P",  10, PCT0,  "c"),
    ("Notice",         "Q",  24, None,  "l"),
    ("Days",           "S",   8, INT0,  "c"),
    ("Availability",   "T",  13, None,  "c"),
    ("Preferred loc",  "AA", 13, None,  "l"),
    ("Loc match",      "AD",  9, INT0,  "c"),
    ("Reason to move", "AB", 22, None,  "l"),
]


def build_recruiter_view(wb):
    ws = wb.create_sheet("Recruiter View")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2.2
    for i, (label, sc, w, fmt, al) in enumerate(RV_COLS):
        ws.column_dimensions[GL(2 + i)].width = w
    last = 1 + len(RV_COLS)

    for r in range(1, 6):
        for c in range(1, last + 2):
            ws.cell(row=r, column=c).fill = fill(CANVAS)
            ws.cell(row=r, column=c).font = F()
    heights(ws, {1: 10, 2: 26, 3: 14, 4: 10, 5: 24})
    ws["B2"] = "Recruiter Working View"
    ws["B2"].font = F(17, True, INK)
    ws["B3"] = ("Every candidate in Master Data-2026, enriched and normalised. "
                "Use the column filters to work a slice — sorting and filtering here never "
                "affects the dashboard.")
    ws["B3"].font = F(8.5, False, MUTED)

    for i, (label, sc, w, fmt, al) in enumerate(RV_COLS):
        c = ws.cell(row=5, column=2 + i, value=label.upper())
        c.font = F(7.5, True, INK_SOFT)
        c.fill = fill(NEUTRAL_TINT)
        c.border = Border(bottom=side(LINE), top=side(LINE))
        c.alignment = CENTERW

    for k in range(R0, R1 + 1):
        rr = 6 + (k - R0)
        ws.row_dimensions[rr].height = 16
        for i, (label, sc, w, fmt, al) in enumerate(RV_COLS):
            cell = ws.cell(row=rr, column=2 + i)
            cell.value = '=IF({C}!$B${k}="","",{C}!${sc}${k})'.format(C=CALC, sc=sc, k=k)
            cell.font = F(8.5, False, INK if label == "Candidate" else INK_SOFT)
            cell.fill = fill(SURFACE if (k - R0) % 2 == 0 else "FAFBFC")
            cell.border = bottom_rule(LINE_SOFT)
            if fmt:
                cell.number_format = fmt
            cell.alignment = Alignment(horizontal={"l": "left", "c": "center"}[al],
                                       vertical="center", indent=1 if al == "l" else 0)

    end = 6 + (R1 - R0)
    ws.auto_filter.ref = "B5:{c}{r}".format(c=GL(last), r=end)
    ws.freeze_panes = "C6"

    idx = {label: GL(2 + i) for i, (label, *_x) in enumerate(RV_COLS)}
    stage_rng = "{c}6:{c}{e}".format(c=idx["Stage"], e=end)
    for needle, bg, fg in (("Hire", GOOD_TINT, GOOD), ("Offer", GOOD_TINT, GOOD),
                           ("Interview", ACCENT_TINT, ACCENT_DK),
                           ("Submission", ACCENT_TINT, ACCENT_DK),
                           ("Withdrawn", RISK_TINT, RISK)):
        ws.conditional_formatting.add(stage_rng, FormulaRule(
            formula=['ISNUMBER(SEARCH("{n}",${c}6))'.format(n=needle, c=idx["Stage"])],
            fill=fill(bg), font=F(8.5, needle in ("Hire", "Offer"), fg), stopIfTrue=True))
    ws.conditional_formatting.add(
        "{c}6:{c}{e}".format(c=idx["Exp"], e=end),
        ColorScaleRule(start_type="num", start_value=0, start_color="FFFFFF",
                       end_type="max", end_color="BFE3CC"))
    ws.conditional_formatting.add(
        "{c}6:{c}{e}".format(c=idx["Hike"], e=end),
        ColorScaleRule(start_type="num", start_value=0, start_color="FFFFFF",
                       end_type="max", end_color="F3C4C4"))
    ws.conditional_formatting.add(
        "{c}6:{c}{e}".format(c=idx["Days"], e=end),
        CellIsRule(operator="between", formula=["0", "30"],
                   fill=fill(GOOD_TINT), font=F(8.5, False, GOOD)))
    ws.conditional_formatting.add(
        "{c}6:{c}{e}".format(c=idx["Days"], e=end),
        CellIsRule(operator="greaterThan", formula=["60"],
                   fill=fill(RISK_TINT), font=F(8.5, False, RISK)))
    ws.conditional_formatting.add(
        "{c}6:{c}{e}".format(c=idx["Loc match"], e=end),
        CellIsRule(operator="equal", formula=["1"],
                   fill=fill(GOOD_TINT), font=F(8.5, True, GOOD)))
    for col in ("CCTC", "ECTC", "Exp", "Days", "Stage #", "Loc match", "Hike"):
        ws.conditional_formatting.add(
            "{c}6:{c}{e}".format(c=idx[col], e=end),
            CellIsRule(operator="equal", formula=["0"], font=F(8.5, False, "C9CED6")))
    return ws


# ==========================================================================
# CANDIDATE DETAIL
# ==========================================================================
IDX = "$C$88"


def get(gcol, fallback='"–"'):
    return ('=IF({i}="",{f},IFERROR(INDEX({C}!${g}${a}:${g}${b},{i}),{f}))'
            ).format(i=IDX, C=CALC, g=gcol, f=fallback, a=R0, b=R1)


def getnum(gcol):
    return ('=IF({i}="",0,IFERROR(INDEX({C}!${g}${a}:${g}${b},{i}),0))'
            ).format(i=IDX, C=CALC, g=gcol, a=R0, b=R1)


def raw(gcol):
    """Inline INDEX expression (no leading '='), for embedding in bigger formulas."""
    return 'IFERROR(INDEX({C}!${g}${a}:${g}${b},{i}),0)'.format(C=CALC, g=gcol, i=IDX,
                                                               a=R0, b=R1)


def _panel_rows(ws, r0, c0, c1, items, label_w=4):
    for i, (label, formula, fmt) in enumerate(items):
        rr = r0 + i
        ws.row_dimensions[rr].height = 18
        ws.merge_cells(start_row=rr, start_column=c0 + 1, end_row=rr, end_column=c0 + label_w)
        a = ws.cell(row=rr, column=c0 + 1, value=label)
        a.font = F(8.5, False, MUTED)
        a.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=rr, start_column=c0 + label_w + 1, end_row=rr, end_column=c1 - 1)
        b = ws.cell(row=rr, column=c0 + label_w + 1, value=formula)
        b.font = F(10, True, INK)
        if fmt:
            b.number_format = fmt
        b.alignment = Alignment(horizontal="left", vertical="center")
        for c in range(c0 + 1, c1):
            ws.cell(row=rr, column=c).border = Border(bottom=side(LINE_SOFT))


CHECKS = [
    ("Experience bar",
     '=IF({w}=0,"Not recorded — no detail on file",IF({w}>=MinExperience,'
     '"Meets the bar — "&TEXT({w},"0.0")&" yrs against "&TEXT(MinExperience,"0")&" required",'
     '"Below the bar — "&TEXT({w},"0.0")&" yrs against "&TEXT(MinExperience,"0")&" required"))',
     '=IF({w}=0,0,IF({w}>=MinExperience,1,-1))'),
    ("Seniority",
     '=IF({w}=0,"Not recorded",IF({w}>=SeniorExperience,"Senior / lead calibre",'
     '"Individual-contributor depth"))',
     '=IF({w}=0,0,IF({w}>=SeniorExperience,1,0))'),
    ("Location fit",
     '=IF({ad}=1,"Based in or relocating to "&TargetLocation,'
     '"Outside the target location — currently "&{h})',
     '=IF({ad}=1,1,-1)'),
    ("Availability",
     '=IF({s}<0,"Notice period not recorded",IF({s}=0,"Immediate joiner",'
     'IF({s}<=MaxNoticeDays,"Available in "&TEXT({s},"0")&" days — within window",'
     '"Notice of "&TEXT({s},"0")&" days exceeds the "&TEXT(MaxNoticeDays,"0")&"-day window")))',
     '=IF({s}<0,0,IF({s}<=MaxNoticeDays,1,-1))'),
    ("Salary expectation",
     '=IF(OR({m}=0,{o}=0),"Compensation not fully recorded",'
     'IF({p}<=0.4,"Uplift of "&TEXT({p},"0%")&" — within a normal band",'
     'IF({p}<=0.7,"Uplift of "&TEXT({p},"0%")&" — above the pool average of "'
     '&TEXT({avg},"0%"),"Uplift of "&TEXT({p},"0%")&" — well above the pool average of "'
     '&TEXT({avg},"0%"))))',
     '=IF(OR({m}=0,{o}=0),0,IF({p}<=0.4,1,IF({p}<=0.7,0,-1)))'),
    ("Pipeline progress",
     '=IF({j}=0,"Not in an active stage",IF({j}>=5,"At offer stage or beyond",'
     'IF({j}=4,"In salary discussion",'
     'IF({j}=3,"Cleared screening and reached client interview",'
     'IF({j}=2,"Submitted to the client","At telephonic screening")))))',
     '=IF({j}=0,-1,IF({j}>=2,1,0))'),
    ("Detail on file",
     '=IF({ac}=1,"Full profile matched in the Submission template",'
     '"No submission detail — experience, company and designation unavailable")',
     '=IF({ac}=1,1,0)'),
    ("Stated reason to move",
     '=IF({ab}="","Not recorded",{ab})',
     '=IF({ab}="",0,1)'),
]

NEXT_STEP = (
    '=IF({i}="","Select a candidate above.",'
    'IF({j}=0,"Candidate has exited the process — no action.",'
    'IF({j}=1,"Complete the telephonic screen, capture CTC and notice, then submit.",'
    'IF({j}=2,"With the client for review — chase feedback and hold an interview slot.",'
    'IF({j}=3,"Interview in progress — collect feedback and prepare the offer case.",'
    'IF({j}=4,"In salary discussion — close the number and confirm the joining date.",'
    '"At offer stage — confirm notice, counter-offer risk and joining date."))))))')


def build_candidate_detail(wb):
    ws = wb.create_sheet("Candidate Detail")
    canvas(ws, last_col=20, last_row=95)
    widths(ws, {"A": 2.2, "R": 2.2, "S": 4})
    for i in range(2, 18):
        ws.column_dimensions[GL(i)].width = 9.6
    ws.column_dimensions["S"].hidden = True
    heights(ws, {1: 10, 2: 26, 3: 14, 4: 10, 5: 8, 6: 13, 7: 22, 8: 8, 9: 10})

    ws["B2"] = "Why This Candidate"
    ws["B2"].font = F(19, True, INK); ws.merge_cells("B2:J2")
    ws["B3"] = ("Pick a name — every panel below re-reads that candidate's row. "
                "Nothing here is typed; the assessments are rules applied to recorded data.")
    ws["B3"].font = F(8.5, False, MUTED); ws.merge_cells("B3:P3")
    for c in range(2, 18):
        ws.cell(row=4, column=c).border = Border(bottom=side(LINE))

    # selector
    ws.merge_cells("B6:G6")
    l = ws.cell(row=6, column=2, value="SELECT CANDIDATE")
    l.font = F(7.5, True, MUTED)
    l.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells("B7:G7")
    v = ws.cell(row=7, column=2, value="Yuva Charan Reddy Durgam")
    v.font = F(12, True, ACCENT_DK)
    v.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for c in range(2, 8):
        ws.cell(row=7, column=c).fill = fill(ACCENT_TINT)
        ws.cell(row=7, column=c).border = box("C7D7FB")

    # hero card
    card(ws, 10, 2, 18, 17)
    heights(ws, {10: 8, 11: 28, 12: 16, 13: 8, 14: 12, 15: 20, 16: 8, 17: 6, 18: 6})
    ws.merge_cells("C11:K11")
    n = ws.cell(row=11, column=3, value=get("B", '"Select a candidate"'))
    n.font = F(18, True, INK)
    n.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("C12:K12")
    sub = ws.cell(row=12, column=3)
    sub.value = ('=IF({i}="","",IF({y}="","Company not recorded",{y})&"  ·  "'
                 '&IF({z}="","Designation not recorded",{z}))'
                 ).format(i=IDX, y=raw("Y"), z=raw("Z"))
    sub.font = F(9.5, False, MUTED)
    ws.merge_cells("L11:Q12")
    rk = ws.cell(row=11, column=12)
    rk.value = ('=IF({i}="","",IF({j}=0,"Exited process","Stage "&TEXT({j},"0")&" of 5  ·  "&{s}))'
                ).format(i=IDX, j=raw("J"), s=raw("I"))
    rk.font = F(11, True, ACCENT)
    rk.alignment = Alignment(horizontal="right", vertical="center")

    chips = [("CITY", get("H"), None, 3), ("EXPERIENCE", getnum("W"), '0.0" yrs";;"–"', 7),
             ("EXPECTED CTC", getnum("O"), MONEY, 11), ("RECRUITER", get("C"), None, 15)]
    for label, formula, fmt, c0 in chips:
        ws.merge_cells(start_row=14, start_column=c0, end_row=14, end_column=c0 + 2)
        a = ws.cell(row=14, column=c0, value=label)
        a.font = F(7, True, MUTED)
        ws.merge_cells(start_row=15, start_column=c0, end_row=15, end_column=c0 + 2)
        b = ws.cell(row=15, column=c0, value=formula)
        b.font = F(12, True, INK)
        if fmt:
            b.number_format = fmt

    # two panels
    band_title(ws, 20, "Profile and process", "Straight from the candidate's row.")
    heights(ws, {20: 16, 21: 13, 22: 6})
    card(ws, 23, 2, 36, 9)
    card(ws, 23, 10, 36, 17)
    h = ws.cell(row=24, column=3, value="PIPELINE"); h.font = F(7.5, True, MUTED)
    h = ws.cell(row=24, column=11, value="COMMERCIALS AND AVAILABILITY"); h.font = F(7.5, True, MUTED)
    _panel_rows(ws, 26, 2, 9, [
        ("Requisition", get("F"), None),
        ("Stage", get("I"), None),
        ("Sub-status", get("K"), None),
        ("Date sourced", get("D"), "dd mmm yyyy"),
        ("Recruiter", get("C"), None),
        ("Current location", get("G"), None),
        ("Normalised city", get("H"), None),
        ("Preferred location", get("AA"), None),
    ])
    _panel_rows(ws, 26, 10, 17, [
        ("Total experience", getnum("W"), '0.0" yrs";;"–"'),
        ("Relevant experience", get("X"), None),
        ("Current CTC", getnum("M"), MONEY),
        ("Expected CTC", getnum("O"), MONEY),
        ("Expected uplift", getnum("P"), PCT),
        ("Notice period", get("Q"), None),
        ("Notice in days", getnum("S"), '0;"Not recorded";"Immediate"'),
        ("Availability band", get("T"), None),
    ])

    # assessment checklist
    band_title(ws, 38, "How this candidate reads against the brief",
               "Eight rules applied to what is on file — green passes, amber is unrecorded or "
               "borderline, red is a gap. Change the parameters on Reference and these re-judge.")
    heights(ws, {38: 16, 39: 13, 40: 6})
    card(ws, 41, 2, 51, 17)
    ctx = dict(i=IDX, w=raw("W"), ad=raw("AD"), h=raw("H"), s=raw("S"), m=raw("M"),
               o=raw("O"), p=raw("P"), j=raw("J"), ac=raw("AC"), ab=raw("AB"),
               avg=kpi_ref("Avg expected hike"))
    for i, (label, text_f, verdict_f) in enumerate(CHECKS):
        rr = 42 + i
        ws.row_dimensions[rr].height = 18
        ws.merge_cells(start_row=rr, start_column=3, end_row=rr, end_column=5)
        a = ws.cell(row=rr, column=3, value=label)
        a.font = F(8.5, True, INK_SOFT)
        a.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=rr, start_column=6, end_row=rr, end_column=16)
        b = ws.cell(row=rr, column=6, value=text_f.format(**ctx))
        b.font = F(9.5, False, INK)
        b.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        s = ws.cell(row=rr, column=19, value=verdict_f.format(**ctx))
        s.font = F(8, False, "C9CED6")
        for c in range(3, 17):
            ws.cell(row=rr, column=c).border = Border(bottom=side(LINE_SOFT))
    rng = "F42:P49"
    for val, bg, fg in ((1, GOOD_TINT, GOOD), (0, WARN_TINT, WARN), (-1, RISK_TINT, RISK)):
        ws.conditional_formatting.add(rng, FormulaRule(
            formula=["$S42={0}".format(val)], fill=fill(bg), font=F(9.5, False, fg),
            stopIfTrue=True))

    # next step + evidence
    band_title(ws, 53, "Recommended next step and evidence",
               "The action follows the stage; the evidence is the raw recruiter record.")
    heights(ws, {53: 16, 54: 13, 55: 6})
    card(ws, 56, 2, 60, 17)
    ws.merge_cells("C57:P58")
    a = ws.cell(row=57, column=3, value=NEXT_STEP.format(i=IDX, j=raw("J")))
    a.font = F(11, True, ACCENT_DK)
    a.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    heights(ws, {56: 6, 57: 18, 58: 18, 59: 6, 60: 6})

    card(ws, 62, 2, 78, 17)
    h = ws.cell(row=63, column=3, value="EVIDENCE ON FILE"); h.font = F(7.5, True, MUTED)
    ev = [("Recruiter comment", get("K", '""')),
          ("Additional notes",
           '=IF({i}="","",IFERROR(IF(TRIM(INDEX({S}!$M$2:$M${e},{i}))="","No further notes '
           'recorded",TRIM(INDEX({S}!$M$2:$M${e},{i}))),""))'.format(i=IDX, S=SRC,
                                                                      e=2 + (R1 - R0))),
          ("Stack evidence", get("X", '"Not recorded"')),
          ("Stated reason for moving", get("AB", '"Not recorded"')),
          ("Current company", get("Y", '"Not recorded"')),
          ("Designation", get("Z", '"Not recorded"'))]
    for i, (label, formula) in enumerate(ev):
        rr = 65 + i * 2
        ws.row_dimensions[rr].height = 13
        ws.row_dimensions[rr + 1].height = 17
        ws.merge_cells(start_row=rr, start_column=3, end_row=rr, end_column=16)
        a = ws.cell(row=rr, column=3, value=label.upper())
        a.font = F(7, True, MUTED)
        ws.merge_cells(start_row=rr + 1, start_column=3, end_row=rr + 1, end_column=16)
        b = ws.cell(row=rr + 1, column=3, value=formula)
        b.font = F(9.5, False, INK_SOFT)
        b.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for c in range(3, 17):
            ws.cell(row=rr + 1, column=c).border = Border(bottom=side(LINE_SOFT))

    # diagnostics
    ws["B87"] = "Lookup diagnostics"
    ws["B87"].font = F(7.5, True, MUTED)
    ws["B88"] = "Grid row"
    ws["B88"].font = F(8, False, MUTED)
    ws["C88"] = '=IFERROR(MATCH($B$7,{C}!$B${a}:$B${b},0),"")'.format(C=CALC, a=R0, b=R1)
    ws["C88"].font = F(8, False, MUTED)
    ws["D88"] = ('=IF($C$88="","No match — pick a name from the list",'
                 '"Reading row "&TEXT($C$88+5,"0")&" of Dashboard Calc")')
    ws["D88"].font = F(8, False, MUTED)
    ws.merge_cells("D88:M88")
    return ws
