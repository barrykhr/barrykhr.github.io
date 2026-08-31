# -*- coding: utf-8 -*-
"""Layer 1: Reference (controlled vocabulary + maps) and Dashboard Calc (engine)."""
from openpyxl.utils import get_column_letter as GL
from theme import *

SRC   = "'Master Data-2026'"
SRC25 = "' Master Data-2025'"
SUB   = "'Submission template'"

R0, R1 = 6, 125          # calc grid data rows
SROW0  = 2               # matching source row for calc row 6
NROWS  = R1 - R0 + 1

SUB_R0, SUB_R1 = 17, 300 # submission-detail block (rows 17+ of Submission template)

STAGES = [
    ("Telephonic Screening", 1, "Progress"),
    ("ZN Submission",        2, "Progress"),
    ("ZN Interview",         3, "Progress"),
    ("ZN Salary Discussion", 4, "Progress"),
    ("ZN Offer",             5, "Progress"),
    ("ZN Hire",              6, "Progress"),
    ("Candidate Withdrawn",  0, "Exit"),
]

LOCMAP = [
    ("Velachery", "Chennai"), ("Taramani", "Chennai"), ("Porur", "Chennai"),
    ("OMR", "Chennai"), ("Pallikaranai", "Chennai"), ("KK Nagar", "Chennai"),
    ("Vyasarpadi", "Chennai"), ("Perungudi", "Chennai"), ("Guindy", "Chennai"),
    ("Adyar", "Chennai"), ("Tambaram", "Chennai"), ("Anna Nagar", "Chennai"),
    ("Sholinganallur", "Chennai"), ("Thoraipakkam", "Chennai"), ("Ambattur", "Chennai"),
    ("Chennai", "Chennai"),
    ("Bengaluru", "Bangalore"), ("Bangalore", "Bangalore"),
    ("Hyderabad", "Hyderabad"), ("Coimbatore", "Coimbatore"), ("Madurai", "Madurai"),
    ("Salem", "Salem"), ("Erode", "Erode"), ("Tirupati", "Tirupati"),
    ("Tirunelveli", "Tirunelveli"), ("Ongole", "Ongole"), ("Puducherry", "Puducherry"),
    ("Kovilpatti", "Kovilpatti"), ("Trichy", "Trichy"), ("Mumbai", "Mumbai"),
    ("Pune", "Pune"), ("Noida", "Noida"), ("Delhi", "Delhi"), ("Gurgaon", "Gurgaon"),
    ("Kochi", "Kochi"), ("Kolkata", "Kolkata"),
]
LOC_SLOTS = 30
LOC_ROW0 = 17
LOC_ROW1 = LOC_ROW0 + LOC_SLOTS - 1
PARAM_ROW = LOC_ROW1 + 4      # hiring parameters block
BAND_ROW  = PARAM_ROW + 11    # banding lists

NOTICE_BUCKETS = ["Immediate", "1-30 days", "31-60 days", "60+ days", "Not recorded"]
EXP_BANDS = [("Under 3 yrs", 0, 3), ("3 - 4 yrs", 3, 4), ("4 - 5 yrs", 4, 5),
             ("5 - 7 yrs", 5, 7), ("7 - 9 yrs", 7, 9), ("9 yrs +", 9, 99)]
HIKE_BANDS = [("Under 20%", 0, .20), ("20 - 40%", .20, .40), ("40 - 60%", .40, .60),
              ("60 - 80%", .60, .80), ("80 - 100%", .80, 1.0), ("Over 100%", 1.0, 99)]
MONTHS = ["January", "February", "March", "April", "May", "June", "July",
          "August", "September", "October", "November", "December"]

FUNNEL = [("Total sourced", 0), ("Telephonic screening", 1), ("Client submission", 2),
          ("Client interview", 3), ("Salary discussion", 4), ("Offer", 5), ("Hire", 6)]

# calc-sheet aggregate block anchors
B_STAGE, B_FUNNEL, B_CITY, B_REQ   = 212, 223, 232, 245
B_RECR,  B_NOTICE, B_EXP, B_MONTH  = 257, 276, 285, 295
B_HIKE,  B_LEAD,   B_TOPX, B_KPI   = 311, 321, 337, 352
B_SCAT,  B_DQ                      = 6,   392        # scatter block lives in BD/BE/BF
N_CITY, N_REQ, N_RECR, N_LEAD, N_TOPX, N_SCAT = 9, 8, 15, 12, 12, 40

MONEY  = '0.0" LPA";-0.0" LPA";"–"'
PCT    = '0.0%;-0.0%;"–"'
PCT0   = '0%;-0%;"–"'
NUM1   = '0.0;-0.0;"–"'
INT0   = '#,##0;-#,##0;"–"'
YRS    = '0.0" yrs";-0.0" yrs";"–"'
DAYS   = '0" days";-0" days";"–"'


def _digit(ref, k):
    """TRUE when character k of ref is 0-9 or a decimal point."""
    m = 'MID({r},{k},1)'.format(r=ref, k=k)
    return 'OR(AND({m}>="0",{m}<="9"),{m}=".")'.format(m=m)


def prefix_len(ref, maxlen=6):
    """Length of the leading run of digits / decimal points, as nested IFs.

    Deliberately NOT IFERROR(VALUE(LEFT(..))) — VALUE happily parses things like
    "1 Month" or "15 Day" as a DATE, which silently turns a notice period into a
    five-digit serial number. Measuring the digit run first makes that impossible.
    """
    out = str(maxlen)
    for n in range(maxlen, 0, -1):
        out = 'IF(NOT({d}),{k},{o})'.format(d=_digit(ref, n), k=n - 1, o=out)
    return out


def ladder(ref, maxlen=6):
    return 'IFERROR(VALUE(LEFT({r},{L})),0)'.format(r=ref, L=prefix_len(ref, maxlen))


def leading_number(ref):
    """0 when the cell is blank, percent-based, or has no numeric prefix."""
    return ('=IF({r}="",0,IF(ISNUMBER(SEARCH("%",{r})),0,{lad}))'
            .format(r=ref, lad=ladder(ref)))


# ==========================================================================
# REFERENCE SHEET
# ==========================================================================
def build_reference(wb):
    ws = wb.create_sheet("Reference")
    canvas(ws, last_col=10, last_row=90)
    widths(ws, {"A": 2, "B": 30, "C": 26, "D": 10, "E": 2, "F": 46, "G": 2})

    ws["B2"] = "Reference & Controlled Values"
    ws["B2"].font = F(13, True, INK)
    ws["B3"] = ("Editable lookup layer. Everything downstream reads from here — "
                "extend these lists and the whole dashboard follows.")
    ws["B3"].font = F(8, False, MUTED)

    # ---- pipeline stages -------------------------------------------------
    ws["B5"] = "PIPELINE STAGES"; ws["B5"].font = F(8, True, MUTED)
    hdr = ["Stage (must match Master Data)", "Funnel order", "Type"]
    for i, h in enumerate(hdr):
        c = ws.cell(row=6, column=2 + i, value=h)
        c.font = F(8, True, INK_SOFT); c.fill = fill(NEUTRAL_TINT)
        c.border = bottom_rule(LINE); c.alignment = LEFT if i == 0 else CENTER
    for i, (name, order, typ) in enumerate(STAGES):
        r = 7 + i
        ws.cell(row=r, column=2, value=name).font = F(9, False, INK)
        c = ws.cell(row=r, column=3, value=order); c.font = F(9, True, ACCENT); c.alignment = CENTER
        c = ws.cell(row=r, column=4, value=typ);   c.font = F(9, False, INK_SOFT); c.alignment = CENTER
        for col in (2, 3, 4):
            ws.cell(row=r, column=col).fill = fill(SURFACE)
            ws.cell(row=r, column=col).border = bottom_rule(LINE_SOFT)
    ws["F6"] = ("Funnel order drives every stage KPI, the funnel chart and the pipeline "
                "ranking. 0 = candidate exited the process.")
    ws["F6"].font = F(8, False, MUTED); ws["F6"].alignment = LEFTW
    ws.merge_cells("F6:F9")

    # ---- location map ----------------------------------------------------
    ws["B15"] = "LOCATION NORMALISATION MAP"; ws["B15"].font = F(8, True, MUTED)
    for i, h in enumerate(["Text found in Current Location", "Normalised city", "#"]):
        c = ws.cell(row=16, column=2 + i, value=h)
        c.font = F(8, True, INK_SOFT); c.fill = fill(NEUTRAL_TINT)
        c.border = bottom_rule(LINE); c.alignment = LEFT if i < 2 else CENTER
    for i in range(LOC_SLOTS):
        r = LOC_ROW0 + i
        if i < len(LOCMAP):
            pat, city = LOCMAP[i]
        else:
            pat, city = "@@unused@@", ""
        ws.cell(row=r, column=2, value=pat).font = F(9, False, INK if i < len(LOCMAP) else MUTED)
        ws.cell(row=r, column=3, value=city).font = F(9, False, INK_SOFT)
        c = ws.cell(row=r, column=4, value=i + 1); c.font = F(8, False, MUTED); c.alignment = CENTER
        for col in (2, 3, 4):
            ws.cell(row=r, column=col).fill = fill(SURFACE)
            ws.cell(row=r, column=col).border = bottom_rule(LINE_SOFT)
    ws["F17"] = ("Free-text locations (‘Porur’, ‘Taramani, Chennai’) are matched against this "
                 "list and rolled up to a city. Rows marked @@unused@@ are spare — overwrite "
                 "them to add a locality; nothing else needs changing.")
    ws["F17"].font = F(8, False, MUTED); ws["F17"].alignment = LEFTW
    ws.merge_cells("F17:F22")

    # ---- hiring parameters ----------------------------------------------
    ws["B%d" % (PARAM_ROW - 1)] = "HIRING PARAMETERS"; ws["B%d" % (PARAM_ROW - 1)].font = F(8, True, MUTED)
    params = [("Target / preferred location", "Chennai",
               "Drives the location-match KPI on every view."),
              ("Minimum experience bar (years)", 4,
               "Drives ‘Meets experience bar’. Compared to Total Exp from Submission template."),
              ("Senior experience bar (years)", 8,
               "Used to size the senior / lead segment of the pool."),
              ("Max acceptable notice (days)", 30,
               "Drives the ‘joiner-ready’ KPI.")]
    for i, h in enumerate(["Parameter", "Value", "What it controls"]):
        c = ws.cell(row=PARAM_ROW, column=2 + i, value=h)
        c.font = F(8, True, INK_SOFT); c.fill = fill(NEUTRAL_TINT)
        c.border = bottom_rule(LINE); c.alignment = LEFT if i != 1 else CENTER
    for i, (nm, val, note) in enumerate(params):
        r = PARAM_ROW + 1 + i
        ws.cell(row=r, column=2, value=nm).font = F(9, False, INK)
        c = ws.cell(row=r, column=3, value=val)
        c.font = F(10, True, ACCENT); c.alignment = CENTER; c.fill = fill(ACCENT_TINT)
        c.border = box(LINE)
        ws.cell(row=r, column=2).fill = fill(SURFACE)
        ws.cell(row=r, column=2).border = bottom_rule(LINE_SOFT)
        n = ws.cell(row=r, column=6, value=note)
        n.font = F(8, False, MUTED); n.alignment = LEFTW
    ws["B%d" % (PARAM_ROW + 6)] = "Blue cells are inputs — edit them."
    ws["B%d" % (PARAM_ROW + 6)].font = F(8, True, ACCENT)

    # ---- notice / experience / hike bands -------------------------------
    ws["B%d" % (BAND_ROW - 2)] = "BANDING"; ws["B%d" % (BAND_ROW - 2)].font = F(8, True, MUTED)
    ws["B%d" % (BAND_ROW - 1)] = "Notice buckets"; ws["B%d" % (BAND_ROW - 1)].font = F(8, True, INK_SOFT)
    for i, b in enumerate(NOTICE_BUCKETS):
        ws.cell(row=BAND_ROW + i, column=2, value=b).font = F(9, False, INK_SOFT)
    ws["C%d" % (BAND_ROW - 1)] = "Experience bands"; ws["C%d" % (BAND_ROW - 1)].font = F(8, True, INK_SOFT)
    for i, (lbl, lo, hi) in enumerate(EXP_BANDS):
        ws.cell(row=BAND_ROW + i, column=3, value=lbl).font = F(9, False, INK_SOFT)
    ws["D%d" % (BAND_ROW - 1)] = "Hike bands"; ws["D%d" % (BAND_ROW - 1)].font = F(8, True, INK_SOFT)
    for i, (lbl, lo, hi) in enumerate(HIKE_BANDS):
        ws.cell(row=BAND_ROW + i, column=4, value=lbl).font = F(9, False, INK_SOFT)
    ws.column_dimensions["D"].width = 16
    return ws


# ==========================================================================
# DASHBOARD CALC  — the engine
# ==========================================================================
GRID = [
    ("A", "#",              10),
    ("B", "Candidate",      22),
    ("C", "Recruiter",      12),
    ("D", "Date",           11),
    ("E", "Month",          11),
    ("F", "Requisition",    24),
    ("G", "Location (raw)", 20),
    ("H", "City",           14),
    ("I", "Stage",          20),
    ("J", "Stage order",    10),
    ("K", "Sub-status",     30),
    ("L", "CCTC (raw)",     16),
    ("M", "CCTC",           10),
    ("N", "ECTC (raw)",     18),
    ("O", "ECTC",           10),
    ("P", "Hike %",         10),
    ("Q", "Notice (raw)",   24),
    ("R", "Notice #",       10),
    ("S", "Notice days",    10),
    ("T", "Notice bucket",  14),
    ("U", "Sub. match",     10),
    ("V", "Total exp (raw)",14),
    ("W", "Total exp",      10),
    ("X", "Relevant exp",   24),
    ("Y", "Company",        24),
    ("Z", "Designation",    24),
    ("AA","Preferred loc",  14),
    ("AB","Reason to move", 22),
    ("AC","Enriched",       9),
    ("AD","Loc match",      9),
    ("AE","Include",        9),
    ("AF","Cand seq",       9),
    ("AG","Recr seq",       9),
    ("AH","Req seq",        9),
    ("AI","City seq",       9),
    ("AJ","Key: pipeline",  14),
    ("AK","Key: experience",14),
    ("AL","Key: hike",      14),
    ("AM","Scatter seq",    10),
]


def build_calc(wb):
    ws = wb.create_sheet("Dashboard Calc")
    ws.sheet_view.showGridLines = False
    for col, head, w in GRID:
        ws.column_dimensions[col].width = w

    ws["A1"] = "Dashboard Calculations"
    ws["A1"].font = F(13, True, INK)
    ws["A2"] = ("Working sheet — do not type here. Every cell is a formula reading "
                "'Master Data-2026' and 'Submission template'. Rows 6-125 mirror source rows "
                "2-121, so up to 120 candidates are supported with no edits.")
    ws["A2"].font = F(8, False, MUTED)
    ws["A4"] = "1 · NORMALISED CANDIDATE GRID"
    ws["A4"].font = F(9, True, ACCENT)

    for col, head, w in GRID:
        c = ws[col + "5"]
        c.value = head
        c.font = F(8, True, INK_SOFT)
        c.fill = fill(NEUTRAL_TINT)
        c.border = bottom_rule(LINE)
        c.alignment = CENTERW
    ws.row_dimensions[5].height = 24
    ws.freeze_panes = "C6"

    for r in range(R0, R1 + 1):
        s = r - R0 + SROW0                       # matching source row
        g = lambda col: "${c}{r}".format(c=col, r=r)
        blank = '$B{r}=""'.format(r=r)

        ws["A%d" % r] = r - R0 + 1
        ws["B%d" % r] = '=IF(TRIM({S}!D{s})="","",TRIM({S}!D{s}))'.format(S=SRC, s=s)
        ws["C%d" % r] = '=IF({b},"",TRIM({S}!A{s}))'.format(b=blank, S=SRC, s=s)
        ws["D%d" % r] = '=IF({b},"",IF({S}!B{s}="","",{S}!B{s}))'.format(b=blank, S=SRC, s=s)
        ws["E%d" % r] = '=IF({b},"",TRIM({S}!C{s}))'.format(b=blank, S=SRC, s=s)
        ws["F%d" % r] = '=IF({b},"",TRIM({S}!G{s}))'.format(b=blank, S=SRC, s=s)
        ws["G%d" % r] = '=IF({b},"",TRIM({S}!F{s}))'.format(b=blank, S=SRC, s=s)
        ws["AN%d" % r] = ('=IF({b},0,IF({g}="",0,'
                          'SUMPRODUCT(MAX(ISNUMBER(SEARCH(LocPattern,{g}))*LocIdx))))'
                          ).format(b=blank, g=g("G"))
        ws["H%d" % r] = ('=IF({b},"",IF({g}="","Not recorded",IF({an}=0,"Other",'
                         'INDEX(LocCity,{an}))))').format(b=blank, g=g("G"), an=g("AN"))
        ws["I%d" % r] = '=IF({b},"",TRIM({S}!H{s}))'.format(b=blank, S=SRC, s=s)
        ws["J%d" % r] = ('=IF({b},0,IFERROR(INDEX(StageOrderRef,MATCH({g},StageNameRef,0)),0))'
                         ).format(b=blank, g=g("I"))
        ws["K%d" % r] = '=IF({b},"",TRIM({S}!I{s}))'.format(b=blank, S=SRC, s=s)
        ws["L%d" % r] = '=IF({b},"",TRIM({S}!J{s}))'.format(b=blank, S=SRC, s=s)
        ws["M%d" % r] = leading_number(g("L"))
        ws["N%d" % r] = '=IF({b},"",TRIM({S}!K{s}))'.format(b=blank, S=SRC, s=s)
        ws["O%d" % r] = leading_number(g("N"))
        ws["P%d" % r] = '=IF(AND({m}>0,{o}>0),{o}/{m}-1,0)'.format(m=g("M"), o=g("O"))
        ws["Q%d" % r] = '=IF({b},"",TRIM({S}!L{s}))'.format(b=blank, S=SRC, s=s)
        ws["R%d" % r] = leading_number(g("Q"))
        ws["S%d" % r] = (
            '=IF({q}="",-1,IF(ISNUMBER(SEARCH("immediate",{q})),0,'
            'IF(ISNUMBER(SEARCH("month",{q})),IF({rn}=0,-1,{rn}*30),'
            'IF(ISNUMBER(SEARCH("day",{q})),IF({rn}=0,-1,{rn}),-1))))'
        ).format(q=g("Q"), rn=g("R"))
        ws["T%d" % r] = (
            '=IF({b},"",IF({s}<0,"Not recorded",IF({s}=0,"Immediate",'
            'IF({s}<=30,"1-30 days",IF({s}<=60,"31-60 days","60+ days")))))'
        ).format(b=blank, s=g("S"))
        ws["U%d" % r] = ('=IF({b},"",IFERROR(MATCH({g},{U}!$D${a}:$D${z},0),""))'
                         ).format(b=blank, g=g("B"), U=SUB, a=SUB_R0, z=SUB_R1)
        enr = '{u}=""'.format(u=g("U"))
        for col, subcol in (("V", "I"), ("X", "J"), ("Y", "K"), ("Z", "M"),
                            ("AA", "O"), ("AB", "S")):
            ws["%s%d" % (col, r)] = (
                '=IF({e},"",IFERROR(TRIM(INDEX({U}!${sc}${a}:${sc}${z},{u})),""))'
            ).format(e=enr, U=SUB, sc=subcol, a=SUB_R0, z=SUB_R1, u=g("U"))
        ws["W%d" % r] = leading_number(g("V"))
        ws["AC%d" % r] = '=IF({b},0,IF({e},0,1))'.format(b=blank, e=enr)
        ws["AD%d" % r] = (
            '=IF({b},0,IF(OR({h}=TargetLocation,AND({p}<>"",'
            'ISNUMBER(SEARCH(TargetLocation,{p})))),1,0))'
        ).format(b=blank, h=g("H"), p=g("AA"))
        ws["AE%d" % r] = (
            '=IF({b},0,IF(AND(OR(SelRecruiter="All",{c}=SelRecruiter),'
            'OR(SelCity="All",{h}=SelCity),OR(SelStage="All",{i}=SelStage),'
            'OR(SelRequisition="All",{f}=SelRequisition)),1,0))'
        ).format(b=blank, c=g("C"), h=g("H"), i=g("I"), f=g("F"))
        ws["AF%d" % r] = '=IF({b},"",COUNT($AF$5:AF{p})+1)'.format(b=blank, p=r - 1)
        ws["AG%d" % r] = ('=IF({c}="","",IF(COUNTIF($C${a}:{c},{c})=1,COUNT($AG$5:AG{p})+1,""))'
                          ).format(c=g("C"), a=R0, p=r - 1).replace("$C$%d:$C%d" % (R0, r),
                                                                    "$C$%d:$C%d" % (R0, r))
        ws["AH%d" % r] = ('=IF({f}="","",IF(COUNTIF($F${a}:$F{r},{f})=1,COUNT($AH$5:AH{p})+1,""))'
                          ).format(f=g("F"), a=R0, r=r, p=r - 1)
        ws["AI%d" % r] = ('=IF({h}="","",IF(COUNTIF($H${a}:$H{r},{h})=1,COUNT($AI$5:AI{p})+1,""))'
                          ).format(h=g("H"), a=R0, r=r, p=r - 1)
        ws["AJ%d" % r] = (
            '=IF(OR({b},{j}=0),"",{j}*100000000+IF(ISNUMBER({d}),{d},0)*1000+({cap}-$A{r}))'
        ).format(b=blank, j=g("J"), d=g("D"), cap=NROWS + 50, r=r)
        ws["AK%d" % r] = ('=IF(OR({b},{w}<=0),"",ROUND({w}*10,0)*100000+({cap}-$A{r}))'
                          ).format(b=blank, w=g("W"), cap=NROWS + 50, r=r)
        ws["AL%d" % r] = ('=IF(OR({b},{p}<=0),"",ROUND({p}*10000,0)*1000+({cap}-$A{r}))'
                          ).format(b=blank, p=g("P"), cap=NROWS + 50, r=r)
        ws["AM%d" % r] = ('=IF(AND({w}>0,{o}>0),COUNT($AM$5:AM{p})+1,"")'
                          ).format(w=g("W"), o=g("O"), p=r - 1)

    # fix the AG running-COUNTIF range (built literally for clarity)
    for r in range(R0, R1 + 1):
        ws["AG%d" % r] = ('=IF($C{r}="","",IF(COUNTIF($C${a}:$C{r},$C{r})=1,'
                          'COUNT($AG$5:AG{p})+1,""))').format(r=r, a=R0, p=r - 1)

    # grid cosmetics
    for r in range(R0, R1 + 1):
        band = SURFACE if (r - R0) % 2 == 0 else "FAFBFC"
        for col, _h, _w in GRID:
            c = ws[col + str(r)]
            c.fill = fill(band)
            c.font = F(8, False, INK_SOFT)
            c.border = bottom_rule(LINE_SOFT)
    for col in ("D",):
        for r in range(R0, R1 + 1):
            ws[col + str(r)].number_format = "dd-mmm-yy"
    for col, fmt in (("M", MONEY), ("O", MONEY), ("P", PCT), ("W", YRS), ("S", INT0)):
        for r in range(R0, R1 + 1):
            ws[col + str(r)].number_format = fmt
    return ws
