# -*- coding: utf-8 -*-
"""Read the computed state of a recalculated workbook."""
import sys, json, warnings
warnings.filterwarnings("ignore")
import openpyxl
from layer1 import (FUNNEL, STAGES, B_KPI, B_FUNNEL, B_STAGE, B_CITY, B_NOTICE, B_EXP, B_LEAD,
                    B_RECR, B_REQ, B_HIKE, B_MONTH, B_DQ, N_LEAD, N_CITY)
from layer2 import KPIS
from layer3 import LEAD_TABLE_ROW, MATRIX_ROW


def probe(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    c = wb["Dashboard Calc"]
    d = wb["Recruitment Dashboard"]
    v = wb["Client View"]
    t = wb["Candidate Detail"]
    out = {}
    out["kpi"] = {KPIS[i][0]: c.cell(row=B_KPI + i, column=2).value
                  for i in range(len(KPIS))}
    out["funnel"] = [(c.cell(row=B_FUNNEL + i, column=1).value,
                      c.cell(row=B_FUNNEL + i, column=3).value) for i in range(len(FUNNEL))]
    out["stages"] = [(c.cell(row=B_STAGE + i, column=1).value,
                      c.cell(row=B_STAGE + i, column=4).value) for i in range(len(STAGES) + 1)]
    out["cities"] = [(c.cell(row=B_CITY + i, column=1).value,
                      c.cell(row=B_CITY + i, column=2).value) for i in range(N_CITY)]
    out["notice"] = [(c.cell(row=B_NOTICE + i, column=1).value,
                      c.cell(row=B_NOTICE + i, column=2).value) for i in range(5)]
    out["expband"] = [(c.cell(row=B_EXP + i, column=1).value,
                       c.cell(row=B_EXP + i, column=2).value) for i in range(6)]
    out["hike"] = [(c.cell(row=B_HIKE + i, column=1).value,
                    c.cell(row=B_HIKE + i, column=2).value) for i in range(6)]
    out["recruiters"] = [(c.cell(row=B_RECR + i, column=1).value,
                          c.cell(row=B_RECR + i, column=2).value,
                          c.cell(row=B_RECR + i, column=3).value) for i in range(6)]
    out["reqs"] = [(c.cell(row=B_REQ + i, column=1).value,
                    c.cell(row=B_REQ + i, column=2).value) for i in range(4)]
    out["leaders"] = [(c.cell(row=B_LEAD + i, column=2).value,
                       c.cell(row=B_LEAD + i, column=6).value,
                       c.cell(row=B_LEAD + i, column=5).value) for i in range(N_LEAD)]
    out["dq"] = [(c.cell(row=B_DQ + i, column=1).value,
                  c.cell(row=B_DQ + i, column=2).value) for i in range(6)]
    out["scatter_n"] = sum(1 for r in range(6, 66)
                           if isinstance(c.cell(row=r, column=56).value, (int, float)))
    out["dash_cards"] = [d.cell(row=r, column=cc).value
                         for r in (12, 18, 24, 30) for cc in (2, 6, 10, 14)]
    out["dash_sel"] = d["N7"].value
    out["dash_leaders"] = [d.cell(row=LEAD_TABLE_ROW + 1 + i, column=3).value
                           for i in range(5)]
    out["dash_matrix"] = [[d.cell(row=MATRIX_ROW + 1 + i, column=cc).value
                           for cc in (3, 8, 11, 12, 13, 14)] for i in range(N_LEAD)]
    out["cv_summary"] = [v.cell(row=r, column=3).value for r in (10, 12, 14)]
    out["cv_top"] = [v.cell(row=35 + i, column=3).value for i in range(5)]
    out["cd_name"] = t["C11"].value
    out["cd_checks"] = [t.cell(row=42 + i, column=6).value for i in range(8)]
    out["cd_verdicts"] = [t.cell(row=42 + i, column=19).value for i in range(8)]
    out["cd_next"] = t["C57"].value
    out["filters"] = [c.cell(row=r, column=60).value for r in range(6, 14)]   # BH
    return out


if __name__ == "__main__":
    print(json.dumps(probe(sys.argv[1]), indent=1, default=str))
